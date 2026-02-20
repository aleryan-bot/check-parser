"""
Check Parser — Streamlit App v2.0
Upload scanned insurance payment checks (PDF or images) and get back a formatted spreadsheet.
"""
import streamlit as st
import anthropic
import base64
import json
import io
import csv
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_bytes
from PIL import Image, ImageEnhance, ImageFilter

# ─── Page config ───
st.set_page_config(page_title="Check Parser", page_icon="🏦", layout="centered")

# ─── Sidebar: API key ───
with st.sidebar:
    st.header("Settings")
    api_key = st.text_input("Anthropic API Key", type="password", help="Get yours at console.anthropic.com → API Keys")
    st.markdown("---")
    st.caption("Check Parser v2.0")
    st.caption("Powered by Claude AI")

# ─── Main UI ───
st.title("Check Parser")
st.markdown("Upload scanned checks (PDF or images) and download a formatted spreadsheet.")

uploaded_files = st.file_uploader(
    "Drop your scanned checks here",
    type=["pdf", "png", "jpg", "jpeg", "tiff", "tif"],
    accept_multiple_files=True,
    help="You can upload a multi-page PDF or multiple image files"
)

EXTRACTION_PROMPT = """You are parsing a scanned insurance payment check image. Extract the data with extreme precision.

IMPORTANT: Read every character carefully. Do NOT guess or approximate. If you cannot read a digit clearly, look at the MICR line (bottom of check) and the written amount line to cross-verify.

STEP-BY-STEP PROCESS:
1. PAYER: Read the company name from the top of the check (logo area or header text).
   - For BlueCross BlueShield of Florida checks from Jacksonville FL:
     * If "Health Options" appears in header → "BlueCross BlueShield of Florida (Health Options)"
     * If "State Employees' PPO Plan" appears → "BCBS FL - State Employees' PPO Plan"
     * Otherwise → "BlueCross BlueShield of Florida"
   - For other BCBS plans, use the full state name: "BlueCross BlueShield of [State]"
   - For other insurers, use the full official name as printed

2. DATE: Look for "Date", "DATE PAID", "ISSUE DATE", or a date box (MO|DAY|YEAR).
   - Output in MM/DD/YYYY format exactly
   - Read each digit individually — do not confuse 0/6/8 or 1/7

3. AMOUNT:
   - Find the NUMERIC amount (usually preceded by $ or asterisks like ******$)
   - Find the WRITTEN amount (e.g. "THREE HUNDRED FOUR & 94/100")
   - BOTH must match. If they don't, re-read more carefully.
   - Output as a number: 304.94 (not a string, no $ sign)

4. BANK: Read the bank name, usually at the bottom or in small text.
   - Common: "Citibank Delaware" (BCBS FL checks), "Regions Bank" (BCBS TN), etc.

5. CHECK NUMBER: Usually top-right or labeled "CHECK NUMBER" or "CHECK NO."
   - Preserve ALL digits including leading zeros
   - Cross-verify with the last group of numbers in the MICR line at the bottom

6. ACCOUNT: Look for "ACCOUNT ID", "PAYEE NUMBER", or similar field on the check face.
   - If none exists, use empty string ""

7. ROUTING: From the MICR line at the very bottom of the check (the line with special banking font).
   - First group of ~9 digits = routing number
   - If not legible, use empty string ""

8. CLAIM: Look for "CLAIM NUMBER" or similar field.
   - If none exists, use empty string ""

Return ONLY a valid JSON object with NO other text, NO markdown, NO explanation:
{"Payer": "", "Date": "", "Amount": 0.00, "Bank": "", "Check_Number": "", "Account": "", "Routing": "", "Claim": ""}"""


def enhance_image(img):
    """Enhance scanned check image for better OCR accuracy."""
    # Convert to RGB if needed
    if img.mode != "RGB":
        img = img.convert("RGB")
    # Increase contrast
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.5)
    # Sharpen
    enhancer = ImageEnhance.Sharpness(img)
    img = enhancer.enhance(2.0)
    return img


def image_to_base64(img):
    """Convert a PIL image to base64-encoded PNG."""
    img = enhance_image(img)
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def bytes_to_base64(data, mime="image/png"):
    """Convert raw bytes to base64, with enhancement for images."""
    try:
        img = Image.open(io.BytesIO(data))
        return image_to_base64(img), "image/png"
    except Exception:
        return base64.standard_b64encode(data).decode("utf-8"), mime


def get_images_from_uploads(files):
    """Convert uploaded files into a list of (base64_data, media_type) tuples."""
    images = []
    for f in files:
        raw = f.read()
        if f.type == "application/pdf":
            # Convert each PDF page to an image at HIGH resolution
            pages = convert_from_bytes(raw, dpi=300)
            for page in pages:
                images.append((image_to_base64(page), "image/png"))
        else:
            b64, mime = bytes_to_base64(raw, f.type if f.type else "image/png")
            images.append((b64, mime))
    return images


def parse_check(client, img_b64, media_type):
    """Send a single check image to Claude and get structured data back."""
    response = client.messages.create(
        model="claude-sonnet-4-5-20250929",
        max_tokens=2048,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": img_b64,
                    },
                },
                {"type": "text", "text": EXTRACTION_PROMPT},
            ],
        }],
    )
    text = response.content[0].text.strip()
    # Clean up any markdown fencing
    if text.startswith("```"):
        text = text.split("\n", 1)[1]
        text = text.rsplit("```", 1)[0]
        text = text.strip()
    # Find JSON object in response
    match = re.search(r'\{[^{}]*\}', text, re.DOTALL)
    if match:
        text = match.group(0)
    data = json.loads(text)
    # Ensure Amount is a float
    if isinstance(data.get("Amount"), str):
        data["Amount"] = float(data["Amount"].replace(",", "").replace("$", ""))
    return data


def generate_xlsx(checks):
    """Generate a formatted XLSX workbook in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Check Register"

    headers = ["#", "Payer", "Date", "Amount", "Bank", "Check Number", "Account #", "Routing #", "Claim #"]
    col_widths = [4, 45, 12, 13, 35, 16, 12, 12, 20]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )
    money_fmt = "$#,##0.00"

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = hdr_font, hdr_fill, border
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, ck in enumerate(checks, 1):
        r = i + 1
        vals = [
            i, ck.get("Payer", ""), ck.get("Date", ""), ck.get("Amount", 0),
            ck.get("Bank", ""), ck.get("Check_Number", ""), ck.get("Account", ""),
            ck.get("Routing", ""), ck.get("Claim", ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border
            if i % 2 == 0:
                c.fill = alt_fill
        ws.cell(row=r, column=4).number_format = money_fmt
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")

    tr = len(checks) + 2
    ws.cell(row=tr, column=3, value="TOTAL").font = Font(bold=True, name="Arial", size=10)
    t = ws.cell(row=tr, column=4, value=f"=SUM(D2:D{tr - 1})")
    t.number_format = money_fmt
    t.font = Font(bold=True, name="Arial", size=10)
    t.border = border

    for i, w in enumerate(col_widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    ws.auto_filter.ref = f"A1:I{tr - 1}"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_csv(checks):
    """Generate a CSV in memory."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["#", "Payer", "Date", "Amount", "Bank", "Check Number", "Account #", "Routing #", "Claim #"])
    for i, ck in enumerate(checks, 1):
        w.writerow([
            i, ck.get("Payer", ""), ck.get("Date", ""),
            f'{ck.get("Amount", 0):.2f}', ck.get("Bank", ""),
            ck.get("Check_Number", ""), ck.get("Account", ""),
            ck.get("Routing", ""), ck.get("Claim", ""),
        ])
    total = sum(c.get("Amount", 0) for c in checks)
    w.writerow(["", "", "TOTAL", f"{total:.2f}", "", "", "", "", ""])
    return buf.getvalue()


# ─── Process ───
if uploaded_files and api_key:
    if st.button("Parse Checks", type="primary", use_container_width=True):
        client = anthropic.Anthropic(api_key=api_key)

        with st.spinner("Converting files to images..."):
            images = get_images_from_uploads(uploaded_files)

        st.info(f"Found {len(images)} check(s) to parse.")

        checks = []
        progress = st.progress(0, text="Parsing checks...")

        for idx, (img_b64, media_type) in enumerate(images):
            progress.progress(idx / len(images), text=f"Parsing check {idx + 1} of {len(images)}...")
            try:
                check_data = parse_check(client, img_b64, media_type)
                checks.append(check_data)
            except Exception as e:
                st.error(f"Error parsing check {idx + 1}: {e}")

        progress.progress(1.0, text="Done!")

        if checks:
            # Show results table
            st.subheader(f"Parsed {len(checks)} check(s)")
            total = sum(c.get("Amount", 0) for c in checks)
            st.metric("Total Amount", f"${total:,.2f}")

            # Display full details table
            display_data = []
            for i, ck in enumerate(checks, 1):
                display_data.append({
                    "#": i,
                    "Payer": ck.get("Payer", ""),
                    "Date": ck.get("Date", ""),
                    "Amount": f"${ck.get('Amount', 0):,.2f}",
                    "Bank": ck.get("Bank", ""),
                    "Check #": ck.get("Check_Number", ""),
                })
            st.table(display_data)

            # Download buttons
            col1, col2 = st.columns(2)
            xlsx_buf = generate_xlsx(checks)
            with col1:
                st.download_button(
                    "Download XLSX",
                    data=xlsx_buf,
                    file_name="check_register.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            csv_data = generate_csv(checks)
            with col2:
                st.download_button(
                    "Download CSV",
                    data=csv_data,
                    file_name="check_register.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

elif uploaded_files and not api_key:
    st.warning("Please enter your Anthropic API key in the sidebar to continue.")
elif not uploaded_files:
    st.info("Upload scanned checks to get started. Supports PDF (single or multi-page), PNG, JPG, and TIFF.")
